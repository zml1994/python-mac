import openpyxl
import sqlite3
from sqlite封装 import get_db_conn, close_db_conn

db_file1 = 'xssmz.db'
wb2 = openpyxl.load_workbook('xsProject.xlsx')
ws = wb2['result 1']
rows = ws.max_row
# 1718
# rows = 700
def main():
    k = 0
    m = 0
    n = 0
    # id = '1a67e64a-2d65-eb11-bfac-8bddac6951'
    # list3 = [('2', id)]
    # up_key(list3, db_file1)
    for i in range(2, rows + 1):
        xsid = ws.cell(i, 12).value
        hzid = ws.cell(i, 13).value
        # print(xsid)
        # print(hzid)
        if hzid == None:
            list3 = [('', 0, xsid)]
            up_key(list3, db_file1)
            n += 1
        elif hzid == 3:
            list3 = [('', 3, xsid)]
            up_key(list3, db_file1)
            m += 1
        else:
            list3 = [(hzid, 1, xsid)]
            up_key(list3, db_file1)
            k += 1
    print(n)
    print(m)
    print(k)


def updata():
    for i in range(2, rows+1):
        core_list = []
        xsid = ws.cell(i, 12).value
        hzid = ws.cell(i, 13).value
        print(xsid)
        print(hzid)


def select_score_all(db_file, id):
    sql = '''select xsProjectCodeis from Project where ProjectID =?
    '''
    conn = sqlite3.connect(db_file)  # 打开或创建数据库文件
    cur = conn.cursor()  # 创建游标
    cur.execute(sql, id)  # 执行sql语句
    # print(cur.fetchall())  # 打印所有结果
    aaa = cur.fetchall()
    cur.close()  # 关闭游标
    conn.close()  # 关闭数据库
    return aaa


def up_key(list3, db_file1):

    sql = '''
    UPDATE Project SET hzProjectCode = ?, hzProjectCodeis= ? WHERE ProjectID = ?
    '''
    conn = get_db_conn(db_file1)
    cur = conn.cursor()  # 创建游标
    cur.executemany(sql, list3)  # 执行sql语句
    conn.commit()  # 提交结果
    close_db_conn(cur, conn)
    return cur.rowcount




if __name__ == "__main__":  # 当程序执行时
    main()